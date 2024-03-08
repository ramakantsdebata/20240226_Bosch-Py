from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def Creating():
    '''Create a new Workbook'''
    wb = Workbook()
    print(wb.sheetnames)
    ws = wb.active
    if ws.title != "Grades":
        ws.title = "Grades"

    wb.save("Trial_01.xlsx")

def UseExisting():
    '''Working with an existing workbook'''
    wb = load_workbook("Trial_01.xlsx")
    ws = wb.active
    print(ws['A1'].value)
    print(ws['A2'].value)
    print(ws['B1'].value)

    ws['A1'].value = "NewString"
    ws['A2'] = "NewData"

    ws.append(['1', '2', '3', '4'])
    ws.append(['1', '2', '3', '4'])
    ws.append(['1', '2', '3', '4'])

    for row in range(1, 11):
        for col in range (1, 6):
            clChar = get_column_letter(col)
            idx = clChar + str(row)
            ws[idx] = "MyData"

    # ws.merge_cells('A1:E1')
    # ws.merge_cells('A1:E10')

    ws.move_range("D4:G6", rows= 2, cols=-3)

    for cell in ws[1:1]:
        cell.value = cell.value.capitalize()

    formula = f"=AVERAGE({col}{row}:{col}{len(data) + 1})"
    ws['B10'] = formula


    wb.save("Trial_01.xlsx")

if __name__ == "__main__":
    # Creating()
    UseExisting()